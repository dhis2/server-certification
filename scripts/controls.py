#!/usr/bin/env python3
import argparse
import html
import os
import re
from typing import Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

ALIGN_RE = re.compile(r"\s*:?[-]+:?\s*")
HEADING_RE = re.compile(r"^###\s+(?:\d+\.[\s-]*)?(.+)$")
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
HEADER_FONT = Font(bold=True)


def clean_cell(text: str) -> str:
    s = html.unescape(text or "").strip().replace("\u00A0", " ")
    s = s.replace("**", "").replace("*", "").replace("`", "")
    return re.sub(r"\s+", " ", s)


def is_table_line(line: str) -> bool:
    t = line.strip()
    return t.startswith("|") and t.endswith("|") and t.count("|") >= 2


def split_md_row(row: str) -> List[str]:
    inner = row.strip()
    if inner.startswith("|"):
        inner = inner[1:]
    if inner.endswith("|"):
        inner = inner[:-1]
    return [clean_cell(p) for p in inner.split("|")]


def is_alignment_row(cells: List[str]) -> bool:
    return bool(cells) and all(ALIGN_RE.fullmatch(c or "") for c in cells)


def parse_md_table(lines: List[str], start_idx: int) -> Tuple[List[List[str]], int]:
    i = start_idx
    block: List[str] = []
    while i < len(lines) and is_table_line(lines[i]):
        block.append(lines[i])
        i += 1

    if not block:
        return [], start_idx

    rows = [split_md_row(r) for r in block]
    if len(rows) >= 2 and is_alignment_row(rows[1]):
        return [rows[0]] + rows[2:], i
    return rows, i


def main() -> None:
    repo_root = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir))
    default_input = os.path.join(repo_root, "controls.md")
    default_output = os.path.join(repo_root, "controls.xlsx")

    parser = argparse.ArgumentParser(description="Convert controls.md to single-sheet Excel checklist.")
    parser.add_argument("--input", "-i", default=default_input, help="Input Markdown file")
    parser.add_argument("--output", "-o", default=default_output, help="Output .xlsx file")
    args = parser.parse_args()

    with open(args.input, "r", encoding="utf-8") as f:
        lines = f.read().splitlines()

    wb = Workbook()
    ws = wb.active
    ws.title = "Controls Checklist"
    ws.freeze_panes = "A2"

    control_cols = {"Control ID", "Control Name", "Type", "IG Level", "CIS v8 Mapping"}
    headers = ["Domain", "Control ID", "Control Name", "Description", "Verification Method", 
               "Type", "IG Level", "CIS v8 Mapping", "Assessment Result", "Evidence/Findings", 
               "Remediation Required"]
    
    widths: Dict[int, int] = {}
    current_domain = ""
    current_row = 1
    i = 0

    # Write headers
    for c_idx, h in enumerate(headers):
        cell = ws.cell(row=current_row, column=c_idx + 1, value=h)
        cell.font = HEADER_FONT
        cell.alignment = WRAP_ALIGN
        widths[c_idx + 1] = max(widths.get(c_idx + 1, 0), len(h) + 2)
    current_row += 1

    while i < len(lines):
        line = lines[i]

        m = HEADING_RE.match(line.strip())
        if m:
            current_domain = clean_cell(m.group(1))
            i += 1
            continue

        if is_table_line(line):
            table, i = parse_md_table(lines, i)
            if not table or len(table) < 2:
                continue

            header = [h.strip() for h in table[0]]
            header_set = set(header)

            if control_cols.issubset(header_set):
                # Extract control data rows
                idx_map = {col: header.index(col) for col in headers[1:] if col in header}
                
                for data_row in table[1:]:
                    row_data = [current_domain]
                    for col in headers[1:]:
                        val = data_row[idx_map[col]] if col in idx_map and idx_map[col] < len(data_row) else ""
                        row_data.append(val)
                    
                    for c_idx, val in enumerate(row_data):
                        cell = ws.cell(row=current_row, column=c_idx + 1, value=val)
                        cell.alignment = WRAP_ALIGN
                        widths[c_idx + 1] = max(widths.get(c_idx + 1, 0), min(len(str(val)) + 2, 60))
                    current_row += 1
            continue

        i += 1

    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    os.makedirs(os.path.dirname(args.output), exist_ok=True)
    wb.save(args.output)


if __name__ == "__main__":
    main()


